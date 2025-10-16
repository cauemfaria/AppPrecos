"""
NFCe Crawler - ULTIMATE VERSION
Fastest possible extraction - only 2 clicks needed!
Uses exact HTML selectors, no tab navigation required
"""

import sys
import io

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import re
import time

def extract_ncm_codes(url, headless=False):
    """Extract NCM codes from NFCe URL - Returns list of products with NCM codes"""
    
    ncm_data = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = browser.new_page()
        
        try:
            # Step 1: Load page
            print("  [1/2] Loading NFCe page...")
            page.goto(url, wait_until="networkidle")
            time.sleep(2)
            
            # Step 2: Click "Visualizar em Abas" (loads ALL NCM data)
            print("  [2/2] Activating tabs view...")
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1)
            page.click('#btnVisualizarAbas')
            time.sleep(4)
            page.wait_for_load_state("networkidle")
            
            # Step 3: Extract from HTML (all NCM codes now loaded!)
            print("  [3/3] Extracting NCM codes from HTML...")
            time.sleep(2)
            
            html = page.content()
            
            # Find all NCM codes with regex
            ncm_matches = re.findall(r'Código NCM</label>\s*<span>(\d{8})</span>', html)
            
            # Find all product names
            product_matches = re.findall(
                r'<table class="toggle box">.*?'
                r'class="fixo-prod-serv-descricao">\s*<span>([^<]+)</span>',
                html,
                re.DOTALL
            )
            
            # Combine products with NCM codes
            for i, (product, ncm) in enumerate(zip(product_matches, ncm_matches), 1):
                ncm_data.append({
                    'number': i,
                    'product': product.strip(),
                    'ncm': ncm
                })
            
            print(f"  ✓ Extracted {len(ncm_data)} products with NCM codes")
            
        except Exception as e:
            print(f"  ✗ Error: {e}")
            
        finally:
            browser.close()
    
    return ncm_data

def save_to_excel(ncm_data, filename="ncm_codes.xlsx"):
    """Save NCM data to formatted Excel file"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "NCM Codes"
    
    # Headers with styling
    headers = ["Nº", "Produto", "Código NCM", "Data Extração"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    extraction_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for idx, item in enumerate(ncm_data, start=2):
        ws[f'A{idx}'] = item['number']
        ws[f'B{idx}'] = item['product']
        ws[f'C{idx}'] = item['ncm']
        ws[f'D{idx}'] = extraction_date
        
        ws[f'A{idx}'].alignment = Alignment(horizontal="center")
        ws[f'C{idx}'].alignment = Alignment(horizontal="center")
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(ncm_data)+1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Save
    wb.save(filename)
    return filename

def main():
    url = "https://www.nfce.fazenda.sp.gov.br/NFCeConsultaPublica/Paginas/ConsultaQRCode.aspx?p=35250948093892001030653080000310101000606075%7C2%7C1%7C1%7Ca9ee07ab1d3a169800dc587738bc26ef7ffbc8db"
    
    print("\n" + "=" * 77)
    print(" NFCe NCM Code Extractor - ULTIMATE OPTIMIZED VERSION")
    print("=" * 77)
    print("\nOptimized for speed - Only essential clicks!")
    print("-" * 77)
    
    start_time = time.time()
    
    # Extract NCM codes
    ncm_data = extract_ncm_codes(url, headless=False)
    
    if len(ncm_data) > 0:
        # Display results
        print("\n  Extracted Products:")
        print("  " + "-" * 73)
        for item in ncm_data:
            prod_short = (item['product'][:42] + "...") if len(item['product']) > 45 else item['product']
            print(f"  [{item['number']:2d}] {prod_short:45s} | NCM: {item['ncm']}")
        print("  " + "-" * 73)
        
        # Save to Excel
        print("\n  Saving to Excel...")
        filename = save_to_excel(ncm_data)
        
        elapsed = time.time() - start_time
        
        print("\n" + "=" * 77)
        print(f" ✓ SUCCESS! Extracted {len(ncm_data)} products in {elapsed:.1f} seconds!")
        print("=" * 77)
        print(f"\n📊 File: {filename}")
        print(f"📦 Products: {len(ncm_data)}")
        print(f"⚡ Speed: {elapsed:.1f} seconds")
        print(f"🕒 Time: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        print("\n✅ Open the Excel file to see all NCM codes!")
    else:
        print("\n❌ No data extracted")
    
    print()

if __name__ == "__main__":
    main()

